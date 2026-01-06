import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import os
import sys

# ======================================================================
# 1. 파일 찾기 함수 (EXE / 개발환경 모두 지원)
# ======================================================================
TEMPLATE_FILE = "rx_test.xlsm"

def find_data_file(filename):
    if getattr(sys, "frozen", False):
        # EXE 환경 (PyInstaller)
        base_path = sys._MEIPASS
    else:
        # 개발 환경
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, filename)

# ======================================================================
# 2. 드롭다운 옵션 로드
# ======================================================================
@st.cache_data
def load_options():
    try:
        template_path = find_data_file(TEMPLATE_FILE)
        wb = openpyxl.load_workbook(template_path, data_only=True, keep_vba=True)
        ws = wb['DATA']

        options = {'J': [], 'K': [], 'L': [], 'M': []}
        col_map = {1:'J', 2:'K', 3:'L', 4:'M'}

        for col_idx, key in col_map.items():
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                if row[0] is not None:
                    options[key].append(str(row[0]))
        return options
    except Exception as e:
        st.error(f"드롭다운 데이터 로드 오류: {e}")
        return {}

# ======================================================================
# 3. 주문 파일 생성
# ======================================================================
ALL_COLUMNS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S']
DROPDOWN_COLS = ['J','K','L','M']

USER_HEADERS = [
    'Order #', 'Eye (R/L)', 'Sph', 'Cyl', 'Axis', 'Prism', 'Add', 'PD', 'HT',
    'Material', 'Products', 'Tint', 'Coating','A', 'B', 'ED', 'DBL', 'Qty', 'Note'
]
HEADER_MAPPING = {f'Col_{ALL_COLUMNS[i]}': USER_HEADERS[i] for i in range(len(ALL_COLUMNS))}

def create_order_file(user_df):
    template_path = find_data_file(TEMPLATE_FILE)

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['ORDER']

    def extract_value(data):
        if isinstance(data, list) and len(data) > 0:
            return data[0]
        return data

    start_row = 3
    for i, row in user_df.iterrows():
        excel_row = start_row + i
        if excel_row > 33:
            break
        for col_index, col_name in enumerate(ALL_COLUMNS):
            df_key = f"Col_{col_name}"
            ws.cell(row=excel_row, column=col_index + 1).value = extract_value(row[df_key])

    # Column Widths
    COLUMN_WIDTHS = {
        'A': 45, 'B': 5, 'C': 5, 'D': 5, 'E': 5, 'F': 8, 'G': 5, 'H': 5, 'I': 5,
        'J': 20, 'K': 50, 'L': 20, 'M': 15, 'N': 20, 'O': 5, 'P': 5, 'Q': 5, 'R': 5, 'S': 5
    }
    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ======================================================================
# 4. Streamlit UI
# ======================================================================
st.set_page_config(page_title="Plazma Order System", layout="wide")

# 타이틀과 이미지를 위한 두 개의 열 생성
col1, col2 = st.columns([0.7, 0.3]) # 70%는 텍스트, 30%는 이미지 공간 할당

with col1:
    st.title("Abba Optical RX Order App")
    st.caption("1. Fill out each line as needed. Per eye (R and L). \n 2. When finished, generate the Excel file at the bottom and download it. \n 3. Email the Excel file to order@abbaopticalusa.com. \n\n *Use the Notes column for any special requests (e.g., Mirror Tint, Base Curve, etc.) \n FT28 availability is below \n CR39 \n CLEAR \n 1.56 \n CLEAR \n BLUE STOP \n PHOTO (GREY, BROWN) \n POLY \n CLEAR")

opts = load_options()
if not opts:
    st.warning("엑셀 템플릿에서 옵션을 불러올 수 없습니다.")
    st.stop()

if 'df_input' not in st.session_state:
    st.session_state.df_input = pd.DataFrame({f'Col_{col}':[None]*31 for col in ALL_COLUMNS})

col_conf = {}
for col in ALL_COLUMNS:
    key = f'Col_{col}'
    header = HEADER_MAPPING[key]
    if col in DROPDOWN_COLS:
        col_conf[key] = st.column_config.SelectboxColumn(header, options=opts[col])
    else:
        col_conf[key] = st.column_config.TextColumn(header)

st.write("### Order Input")
edited_df = st.data_editor(
    st.session_state.df_input,
    column_config=col_conf,
    num_rows="fixed",
    hide_index=False,
    height=600,
    use_container_width=True
)

st.write("---")

if st.button("GENERATE EXCEL / 🚀 엑셀 파일 생성 및 다운로드", type="primary"):
    excel_file = create_order_file(edited_df)
    st.download_button(
        label="📥 DOWNLOAD (.xlsm)",
        data=excel_file,
        file_name="Plazma_Order.xlsm",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success("DONE! DOWNLOAD AND CHECK YOUR EXCEL. 완료! 다운로드된 엑셀을 확인하세요.")
